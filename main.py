import openpyxl
import xlrd

file_path = 'Data K2.xlsx'
workbook = openpyxl.load_workbook(file_path, data_only=True)
sheet = workbook.active

def get_cell_value(sheet, row, col):
    excel_value = sheet.cell(row=row, column=col).value
    if excel_value and isinstance(excel_value, str) and excel_value.startswith('='):
        rb = xlrd.open_workbook(file_path, formatting_info=True)
        sheet_xlrd = rb.sheet_by_index(0)
        cell_value = sheet_xlrd.cell_value(row - 1, col - 1)
        return cell_value
    else:
        return excel_value

max_revenue = 0
product_with_max_revenue = ''
for row in range(2, sheet.max_row + 1):
    revenue = get_cell_value(sheet, row, 7)
    if isinstance(revenue, (float, int)) and revenue > max_revenue:
        max_revenue = revenue
        product_with_max_revenue = sheet.cell(row=row, column=1).value

print(f"Продукт с наибольшей выручкой: {product_with_max_revenue}")


total_revenue = sum(sheet.cell(row=row, column=7).value or 0 for row in range(2, sheet.max_row + 1))
print(f"Общая сумма выручки: {total_revenue}")


total_sales_price = sum(sheet.cell(row=row, column=7).value or 0 for row in range(2, sheet.max_row + 1))
total_quantity = sum(sheet.cell(row=row, column=4).value or 0 for row in range(2, sheet.max_row + 1))

if total_quantity != 0:
    average_sales_price = total_sales_price / total_quantity
    print(f"Средняя стоимость продажи продукта: {average_sales_price:.2f}")
else:
    print("Невозможно рассчитать среднюю стоимость, так как общее количество равно нулю.")

valid_rows = 0
total_profit_percentage = 0

for row in range(2, sheet.max_row + 1):
    revenue = sheet.cell(row=row, column=7).value
    cost_price = sheet.cell(row=row, column=2).value

    if revenue is not None and cost_price is not None:
        profit_percentage = ((revenue - cost_price) / cost_price) * 100
        total_profit_percentage += profit_percentage
        valid_rows += 1

if valid_rows > 0:
    average_profit_percentage = total_profit_percentage / valid_rows
    print(f"Средний процент прибыли: {average_profit_percentage:.2f}%")
else:
    print("Невозможно рассчитать средний процент прибыли из-за отсутствия данных.")

customer_revenue = {}
customer_count = {}

for row in range(2, sheet.max_row + 1):
    revenue = sheet.cell(row=row, column=7).value
    customer_name = sheet.cell(row=row, column=8).value

    if revenue is not None:
        if customer_name not in customer_revenue:
            customer_revenue[customer_name] = 0
            customer_count[customer_name] = 0

        customer_revenue[customer_name] += revenue
        customer_count[customer_name] += 1

for customer, revenue in customer_revenue.items():
    avg_sales_for_customer = revenue / customer_count[customer] if customer_count[customer] != 0 else 0
    print(
        f"Покупатель: {customer}, Общая сумма выручки: {revenue}, Средняя стоимость продажи: {avg_sales_for_customer:.2f}")

average_cost = sum(sheet.cell(row=row, column=2).value or 0 for row in range(2, sheet.max_row + 1)) / (sheet.max_row - 1)
products_above_avg_cost = [
    sheet.cell(row=row, column=1).value
    for row in range(2, sheet.max_row + 1)
    if (sheet.cell(row=row, column=2).value or 0) > average_cost
]
print("Продукты с себестоимостью выше средней:", products_above_avg_cost)

gender_sales = {'Мужской': {'revenue': 0, 'count': 0, 'names': ['Иван', 'Петр', 'Денис']},
                'Женский': {'revenue': 0, 'count': 0, 'names': ['Ольга']}}

male_endings = ['ов', 'ев', 'ин', 'ын', 'ой', 'ий', 'ый', 'ч', 'й']
female_endings = ['ва', 'на', 'ая', 'яя', 'ь', 'ия', 'ая']

for row in range(2, sheet.max_row + 1):
    revenue = sheet.cell(row=row, column=7).value
    customer_name = sheet.cell(row=row, column=8).value

    if customer_name:
        name = customer_name.strip().lower()
        gender = 'Мужской' if any(name.endswith(end) for end in male_endings) else 'Женский' if any(
            name.endswith(end) for end in female_endings) else None

        if gender:
            if revenue is not None:
                gender_sales[gender]['revenue'] += revenue
                gender_sales[gender]['count'] += 1
                gender_sales[gender]['names'].append(customer_name)

for gender, data in gender_sales.items():
    total_revenue = data['revenue']
    count = data['count']
    average_sales = total_revenue / count if count != 0 else 0
    print(f"Пол: {gender}, Общая сумма выручки: {total_revenue}, Средняя стоимость продажи: {average_sales:.2f}")
    print(f"Имена: {data['names']}")


workbook.close()
